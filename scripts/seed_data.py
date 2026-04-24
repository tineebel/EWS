# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
"""
EWS Data Seeder
Reads from:  DocSF/All Employee_masked.xlsx  → Organization + Employee data
             DocSF/flowWF.xlsx               → Workflow Templates
Inserts to:  192.168.0.155 / EWS
"""
import pyodbc
import openpyxl
import uuid
from datetime import datetime, timezone
from collections import defaultdict

# ────────────────────────────────────────────────────────────
# Connection
# ────────────────────────────────────────────────────────────
CONN_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=192.168.0.155;"
    "DATABASE=EWS;"
    "UID=sa;"
    "PWD=SFcinem@;"
    "TrustServerCertificate=yes;"
)

NOW = datetime.now()
CREATED_BY = "SEEDER"

# ────────────────────────────────────────────────────────────
# JobGrade mapping (enum values in C#)
# ────────────────────────────────────────────────────────────
JOB_GRADE_MAP = {
    "XX": 0, "A0": 1, "A1": 2, "A2": 3, "A3": 4,
    "B0": 5, "B1": 6, "B2": 7, "C1": 8, "C3": 8, "D1": 9
}

# WfScopeType: All=0, Branch=1, Ho=2
def get_wf_scope(position_code: str) -> int:
    if position_code.startswith("CB"):
        return 1  # Branch
    if position_code.startswith("HO"):
        return 2  # Ho
    return 0  # All

# IsChiefLevel: grade B0 and above
def is_chief(grade_str: str) -> bool:
    return grade_str in ("A0", "A1", "A2", "A3", "B0")

# ────────────────────────────────────────────────────────────
# WorkflowStep definitions per DocCode
# ApproverType enum: DirectSupervisor=1, SectionManager=2,
#   DeptManager=3, DivisionDirector=4, CLevel=5, CEO=6, SpecificPosition=7
# ────────────────────────────────────────────────────────────
# Steps are defined by condition value ranges
def get_steps_for_condition(doc_code: int, condition1: str, wf_type: str) -> list[dict]:
    """
    Returns list of step dicts: {name, approver_type, specific_code}
    Based on doc_code category + amount condition
    """
    cond = (condition1 or "NULL").strip()

    # ── Memo (1001, 1002) ──────────────────────────────────
    if doc_code == 1001:
        return [
            {"name": "Direct Supervisor", "type": 1},
            {"name": "Department Manager", "type": 3},
        ]
    if doc_code == 1002:
        return [
            {"name": "Direct Supervisor", "type": 1},
            {"name": "Department Manager", "type": 3},
            {"name": "Division Director", "type": 4},
        ]

    # ── Write Off (2101) ───────────────────────────────────
    if doc_code == 2101:
        return [
            {"name": "Department Manager",  "type": 3},
            {"name": "Division Director",   "type": 4},
            {"name": "C-Level",             "type": 5},
            {"name": "CEO",                 "type": 6},
        ]

    # ── IT Requests (4001-4006) ────────────────────────────
    if 4001 <= doc_code <= 4006:
        return [
            {"name": "Direct Supervisor",  "type": 1},
            {"name": "Department Manager", "type": 3},
        ]

    # ── SIS (5001-5008) ────────────────────────────────────
    if 5001 <= doc_code <= 5008:
        steps = [{"name": "Direct Supervisor", "type": 1}]
        if cond not in ("NULL", "", "> 0"):
            steps.append({"name": "Section Manager", "type": 2})
        if _over(cond, 3000):
            steps.append({"name": "Department Manager", "type": 3})
        if _over(cond, 50000):
            steps.append({"name": "Division Director", "type": 4})
        if _over(cond, 300000):
            steps.append({"name": "C-Level", "type": 5})
        if _over(cond, 1000000):
            steps.append({"name": "CEO", "type": 6})
        return steps

    # ── Contract / Customer (6001-6003) ───────────────────
    if 6001 <= doc_code <= 6003:
        return [
            {"name": "Department Manager", "type": 3},
            {"name": "Division Director",  "type": 4},
            {"name": "C-Level",            "type": 5},
            {"name": "CEO",                "type": 6},
        ]

    # ── Finance (2001-2010) — amount-based ────────────────
    steps = [{"name": "Direct Supervisor", "type": 1}]
    if _over(cond, 1000):
        steps.append({"name": "Section Manager", "type": 2})
    if _over(cond, 5000):
        steps.append({"name": "Department Manager", "type": 3})
    if _over(cond, 20000):
        steps.append({"name": "Division Director", "type": 4})
    if _over(cond, 300000):
        steps.append({"name": "C-Level", "type": 5})
    if _over(cond, 1000000):
        steps.append({"name": "CEO", "type": 6})
    return steps


def _over(cond: str, threshold: float) -> bool:
    """True if condition implies amount > threshold"""
    cond = cond.replace(",", "").strip()
    if cond.startswith(">"):
        try:
            val = float(cond.lstrip(">= ").strip())
            return val >= threshold
        except:
            return False
    if cond.startswith("<="):
        try:
            val = float(cond[2:].strip())
            return val >= threshold
        except:
            return False
    return False


# ────────────────────────────────────────────────────────────
# DocCode metadata
# ────────────────────────────────────────────────────────────
DOC_META = {
    1001: ("Memo ทั่วไป", "General Memo", "Memo"),
    1002: ("Memo ผ่าน DOA", "Memo via DOA", "Memo"),
    2001: ("ใบสำคัญจ่าย-สาขา (PCV-BR)", "Petty Cash Voucher Branch", "Finance"),
    2002: ("ใบสำคัญจ่าย-สำนักงาน (PCV-HO)", "Petty Cash Voucher HO", "Finance"),
    2003: ("ใบขอเบิก-สาขา (PCR-BR)", "Petty Cash Request Branch", "Finance"),
    2004: ("ใบขอเบิก-สำนักงาน (PCR-HO)", "Petty Cash Request HO", "Finance"),
    2005: ("เบิกล่วงหน้า-สาขา (ADV-BR)", "Advance Branch", "Finance"),
    2006: ("เบิกล่วงหน้า-สำนักงาน (ADV-HO)", "Advance HO", "Finance"),
    2007: ("เบิกเพิ่ม-สาขา (ADC-BR)", "Additional Advance Branch", "Finance"),
    2008: ("เบิกเพิ่ม-สำนักงาน (ADC-HO)", "Additional Advance HO", "Finance"),
    2009: ("ค่าใช้จ่าย-สาขา (EXP-BR)", "Expense Branch", "Finance"),
    2010: ("ค่าใช้จ่าย-สำนักงาน (EXP-HO)", "Expense HO", "Finance"),
    2101: ("Write Off ทรัพย์สิน", "Asset Write Off", "Finance"),
    4001: ("ขอ Email พนักงานใหม่", "New Employee Email Request", "IT"),
    4002: ("ขอสิทธิ์ VPN", "VPN Access Request", "IT"),
    4003: ("ขอสิทธิ์ Application", "Application Access Request", "IT"),
    4004: ("เปลี่ยนแปลงสิทธิ์ระบบ", "System Permission Change", "IT"),
    4005: ("ขอสิทธิ์ Application กรณีพิเศษ", "Special Application Access", "IT"),
    4006: ("ยกเลิกสิทธิ์", "Access Revocation", "IT"),
    5001: ("SIS: Accrued", "SIS Accrued", "SIS"),
    5002: ("SIS: Forecast & PR", "SIS Forecast & PR", "SIS"),
    5003: ("SIS: PO (สาขา)", "SIS Purchase Order Branch", "SIS"),
    5004: ("SIS: PO (สำนักงาน)", "SIS Purchase Order HO", "SIS"),
    5005: ("SIS: PO D365", "SIS Purchase Order D365", "SIS"),
    5006: ("SIS: Transfer Order", "SIS Transfer Order", "SIS"),
    5007: ("SIS: ยกเลิก Receipt PO", "SIS Cancel Receipt PO", "SIS"),
    5008: ("SIS: รับเงินคืนหลังจ่าย", "SIS Refund After Payment", "SIS"),
    6001: ("สัญญาต่างๆ", "Contracts", "Contract"),
    6002: ("เงินมัดจำ", "Deposit", "Contract"),
    6003: ("ลูกค้า", "Customer", "Contract"),
}

SCOPE_MAP = {"ALL": 0, "All": 0, "Branch": 1, "Ho": 2}


# ────────────────────────────────────────────────────────────
# MAIN SEEDER
# ────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("EWS Data Seeder")
    print("=" * 60)

    conn = pyodbc.connect(CONN_STR, autocommit=False)
    cur = conn.cursor()

    try:
        seed_organization(cur, conn)
        seed_workflow_config(cur, conn)
        print("\n✓ Seeding complete!")
    except Exception as e:
        conn.rollback()
        print(f"\n✗ Error: {e}")
        raise
    finally:
        conn.close()


# ────────────────────────────────────────────────────────────
# ORGANIZATION SEED
# ────────────────────────────────────────────────────────────
def seed_organization(cur, conn):
    print("\n── Organization ──")
    wb = openpyxl.load_workbook("D:/ClaudeCode/EWS/DocSF/All Employee_masked.xlsx")
    ws = wb["v_employee_pdp_masked"]

    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]

    # ── 1. Divisions ──────────────────────────────────────
    divisions = {}  # code → name
    for r in rows:
        code, name = r[15], r[16]
        if code and name and code not in divisions:
            divisions[code] = str(name).strip()

    div_id_map = {}  # code → db id
    for code, name in divisions.items():
        cur.execute("SELECT DivisionId FROM dbo.Divisions WHERE DivisionCode=?", code)
        row = cur.fetchone()
        if row:
            div_id_map[code] = row[0]
        else:
            cur.execute("""
                INSERT INTO dbo.Divisions (DivisionCode,DivisionName,IsActive,CreatedAt,CreatedBy)
                VALUES (?,?,1,?,?)
            """, code, name, NOW, CREATED_BY)
            cur.execute("SELECT @@IDENTITY")
            div_id_map[code] = int(cur.fetchone()[0])

    conn.commit()
    print(f"  Divisions: {len(div_id_map)}")

    # ── 2. Departments ────────────────────────────────────
    departments = {}  # code → (name, div_code)
    for r in rows:
        code, name, div = r[13], r[14], r[15]
        if code and name and div and code not in departments:
            departments[code] = (str(name).strip(), div)

    dept_id_map = {}
    for code, (name, div) in departments.items():
        cur.execute("SELECT DepartmentId FROM dbo.Departments WHERE DeptCode=?", code)
        row = cur.fetchone()
        if row:
            dept_id_map[code] = row[0]
        else:
            cur.execute("""
                INSERT INTO dbo.Departments (DeptCode,DeptName,DivisionId,IsActive,CreatedAt,CreatedBy)
                VALUES (?,?,?,1,?,?)
            """, code, name, div_id_map.get(div, 1), NOW, CREATED_BY)
            cur.execute("SELECT @@IDENTITY")
            dept_id_map[code] = int(cur.fetchone()[0])

    conn.commit()
    print(f"  Departments: {len(dept_id_map)}")

    # ── 3. Sections ───────────────────────────────────────
    sections = {}  # code → (name, dept_code)
    for r in rows:
        code, name, dept = r[11], r[12], r[13]
        if code and name and dept and code not in sections:
            sections[code] = (str(name).strip(), dept)

    sect_id_map = {}
    for code, (name, dept) in sections.items():
        cur.execute("SELECT SectionId FROM dbo.Sections WHERE SectCode=?", code)
        row = cur.fetchone()
        if row:
            sect_id_map[code] = row[0]
        else:
            cur.execute("""
                INSERT INTO dbo.Sections (SectCode,SectName,DepartmentId,IsActive,CreatedAt,CreatedBy)
                VALUES (?,?,?,1,?,?)
            """, code, name, dept_id_map.get(dept, 1), NOW, CREATED_BY)
            cur.execute("SELECT @@IDENTITY")
            sect_id_map[code] = int(cur.fetchone()[0])

    conn.commit()
    print(f"  Sections: {len(sect_id_map)}")

    # ── 4. Positions (Pass 1: insert without parent) ─────
    positions = {}  # code → (name, short, grade, sect, reportto_code)
    for r in rows:
        pos_code = r[7]
        if pos_code and pos_code not in positions:
            positions[pos_code] = {
                "name":     str(r[8]).strip() if r[8] else pos_code,
                "short":    str(r[9]).strip() if r[9] else None,
                "grade":    str(r[10]).strip() if r[10] else "D1",
                "sect":     r[11],
                "reportto": r[19],  # reportto_position_code
            }

    pos_id_map = {}  # code → db id
    for code, p in positions.items():
        cur.execute("SELECT PositionId FROM dbo.Positions WHERE PositionCode=?", code)
        row = cur.fetchone()
        if row:
            pos_id_map[code] = row[0]
        else:
            grade_val = JOB_GRADE_MAP.get(p["grade"], 9)
            scope_val = get_wf_scope(code)
            chief_val = 1 if is_chief(p["grade"]) else 0
            sect_id  = sect_id_map.get(p["sect"], list(sect_id_map.values())[0])
            cur.execute("""
                INSERT INTO dbo.Positions
                (PositionCode,PositionName,PositionShortName,JobGrade,WfScopeType,
                 SectionId,IsChiefLevel,IsActive,CreatedAt,CreatedBy)
                VALUES (?,?,?,?,?,?,?,1,?,?)
            """, code, p["name"], p["short"], grade_val, scope_val,
                 sect_id, chief_val, NOW, CREATED_BY)
            cur.execute("SELECT @@IDENTITY")
            pos_id_map[code] = int(cur.fetchone()[0])

    conn.commit()
    print(f"  Positions: {len(pos_id_map)}")

    # ── 5. Positions (Pass 2: update ParentPositionId) ───
    updated = 0
    for code, p in positions.items():
        reportto = p["reportto"]
        if reportto and reportto in pos_id_map and reportto != code:
            cur.execute("""
                UPDATE dbo.Positions SET ParentPositionId=? WHERE PositionCode=?
            """, pos_id_map[reportto], code)
            updated += 1

    conn.commit()
    print(f"  Position hierarchy links: {updated}")

    # ── 6. Employees ─────────────────────────────────────
    employees = {}  # guid_str → row
    for r in rows:
        guid_str = r[0]
        if guid_str and guid_str not in employees:
            employees[guid_str] = r

    # Bulk insert employees
    cur.execute("SELECT EmployeeId, EmployeeCode FROM dbo.Employees")
    rows_emps = cur.fetchall()
    existing_emps  = {str(row[0]).lower() for row in rows_emps}
    seen_emp_codes = {row[1] for row in rows_emps}

    emp_batch = []
    for guid_str, r in employees.items():
        try:
            emp_guid = str(uuid.UUID(str(guid_str)))
        except:
            continue
        if emp_guid.lower() in existing_emps:
            continue

        status_val = 1 if r[1] == "A" else 0
        start_date = r[23] if isinstance(r[23], datetime) else NOW
        end_date   = r[24] if isinstance(r[24], datetime) else None
        emp_code_raw = str(r[2]).strip() if r[2] else emp_guid[:8]
        name_th  = str(r[3]).strip() if r[3] else ""
        name_en  = str(r[4]).strip() if r[4] else ""
        email    = str(r[21]).strip() if r[21] else None
        tel      = str(r[6]).strip() if r[6] else None
        img      = str(r[22]).strip() if r[22] else None
        is_test  = 1 if r[26] else 0

        emp_code = emp_code_raw
        suffix = 1
        while emp_code in seen_emp_codes:
            emp_code = f"{emp_code_raw}-{suffix}"
            suffix += 1
        seen_emp_codes.add(emp_code)

        emp_batch.append((emp_guid, emp_code, name_th or name_en, name_en,
                          tel, email, img, status_val, start_date, end_date,
                          is_test, NOW, CREATED_BY))

    if emp_batch:
        cur.fast_executemany = True
        cur.executemany("""
            INSERT INTO dbo.Employees
            (EmployeeId,EmployeeCode,EmployeeName,EmployeeNameEn,
             Tel,Email,ImagePath,Status,StartDate,EndDate,IsTest,CreatedAt,CreatedBy)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, emp_batch)

    conn.commit()
    emp_count = len(emp_batch)
    print(f"  Employees: {emp_count}")

    # ── 7. PositionAssignments (bulk) ────────────────────
    cur.execute("SELECT EmployeeId, PositionId FROM dbo.PositionAssignments")
    existing_asgn = {(str(row[0]).lower(), row[1]) for row in cur.fetchall()}

    asgn_batch = []
    for guid_str, r in employees.items():
        try:
            emp_guid = str(uuid.UUID(str(guid_str)))
        except:
            continue
        pos_code = r[7]
        if not pos_code or pos_code not in pos_id_map:
            continue
        pos_id = pos_id_map[pos_code]
        if (emp_guid.lower(), pos_id) in existing_asgn:
            continue

        start_date = r[23] if isinstance(r[23], datetime) else NOW
        end_date   = r[24] if isinstance(r[24], datetime) else None
        is_active  = 1 if r[1] == "A" else 0

        asgn_batch.append((emp_guid, pos_id, start_date, end_date, is_active, NOW, CREATED_BY))

    if asgn_batch:
        cur.fast_executemany = True
        cur.executemany("""
            INSERT INTO dbo.PositionAssignments
            (EmployeeId,PositionId,StartDate,EndDate,IsVacant,IsActive,CreatedAt,CreatedBy)
            VALUES (?,?,?,?,0,?,?,?)
        """, asgn_batch)

    conn.commit()
    asgn_count = len(asgn_batch)
    print(f"  PositionAssignments: {asgn_count}")


# ────────────────────────────────────────────────────────────
# WORKFLOW CONFIG SEED
# ────────────────────────────────────────────────────────────
def seed_workflow_config(cur, conn):
    print("\n── Workflow Config ──")
    wb = openpyxl.load_workbook("D:/ClaudeCode/EWS/DocSF/flowWF.xlsx")
    ws = wb["Sheet1"]

    flow_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                 if r[0] and isinstance(r[0], (int, float))]

    # ── 1. DocumentTypes ─────────────────────────────────
    doc_type_id_map = {}
    doc_codes_seen = set()
    for r in flow_rows:
        dc = int(r[0])
        if dc in doc_codes_seen:
            continue
        doc_codes_seen.add(dc)

        meta = DOC_META.get(dc, (f"Document {dc}", f"Document {dc}", "Other"))
        cur.execute("SELECT DocumentTypeId FROM wf.DocumentTypes WHERE DocCode=?", dc)
        row = cur.fetchone()
        if row:
            doc_type_id_map[dc] = row[0]
        else:
            cur.execute("""
                INSERT INTO wf.DocumentTypes
                (DocCode,DocName,DocNameEn,Category,IsActive,CreatedAt,CreatedBy)
                VALUES (?,?,?,?,1,?,?)
            """, dc, meta[0], meta[1], meta[2], NOW, CREATED_BY)
            cur.execute("SELECT @@IDENTITY")
            doc_type_id_map[dc] = int(cur.fetchone()[0])

    conn.commit()
    print(f"  DocumentTypes: {len(doc_type_id_map)}")

    # ── 2. WorkflowTemplates + Steps ─────────────────────
    tmpl_count = 0
    step_count = 0

    for r in flow_rows:
        dc           = int(r[0])
        flow_code    = int(r[1])
        flow_desc    = str(r[2]).strip() if r[2] else f"Flow {flow_code}"
        wf_type_str  = str(r[4]).strip() if r[4] else "ALL"
        special      = 1 if r[5] == 1.0 else 0
        urgent       = 1 if r[6] == 1.0 else 0
        cond1        = str(r[7]).strip() if r[7] and r[7] != "NULL" else None
        cond2        = str(r[8]).strip() if r[8] and r[8] != "NULL" else None
        cond3        = str(r[9]).strip() if r[9] and r[9] != "NULL" else None
        cond4        = str(r[10]).strip() if r[10] and r[10] != "NULL" else None
        cond5        = str(r[11]).strip() if r[11] and r[11] != "NULL" else None

        scope_val    = SCOPE_MAP.get(wf_type_str, 0)
        doc_type_id  = doc_type_id_map.get(dc)
        if not doc_type_id:
            continue

        # Check existing template
        cur.execute("""
            SELECT TemplateId FROM wf.WorkflowTemplates
            WHERE DocumentTypeId=? AND FlowCode=?
        """, doc_type_id, flow_code)
        row = cur.fetchone()
        if row:
            tmpl_id = row[0]
        else:
            cur.execute("""
                INSERT INTO wf.WorkflowTemplates
                (DocumentTypeId,FlowCode,FlowDesc,WfScopeType,HasSpecialItem,IsUrgent,
                 Condition1,Condition2,Condition3,Condition4,Condition5,
                 IsActive,CreatedAt,CreatedBy)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,1,?,?)
            """, doc_type_id, flow_code, flow_desc, scope_val, special, urgent,
                 cond1, cond2, cond3, cond4, cond5, NOW, CREATED_BY)
            cur.execute("SELECT @@IDENTITY")
            tmpl_id = int(cur.fetchone()[0])
            tmpl_count += 1

            # Insert Steps for this template
            steps = get_steps_for_condition(dc, cond1 or "NULL", wf_type_str)
            for i, step in enumerate(steps, 1):
                cur.execute("""
                    INSERT INTO wf.WorkflowSteps
                    (TemplateId,StepOrder,StepName,ApproverType,SpecificPositionCode,
                     EscalationDays,IsRequired,IsActive,CreatedAt,CreatedBy)
                    VALUES (?,?,?,?,?,?,1,1,?,?)
                """, tmpl_id, i, step["name"],
                     step["type"],
                     step.get("code"),
                     3,  # EscalationDays default = 3 วัน
                     NOW, CREATED_BY)
                step_count += 1

    conn.commit()
    print(f"  WorkflowTemplates: {tmpl_count}")
    print(f"  WorkflowSteps: {step_count}")


if __name__ == "__main__":
    main()
